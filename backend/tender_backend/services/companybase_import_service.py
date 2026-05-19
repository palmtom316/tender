"""Companybase workbook validation and import."""

from __future__ import annotations

import json
import re
import tarfile
import tempfile
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any
from uuid import UUID, uuid4

from openpyxl import load_workbook
from psycopg import Connection
from psycopg.rows import dict_row


DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")
ALLOWED_SUFFIXES = {".pdf", ".png", ".jpg", ".jpeg", ".bmp", ".gif", ".tif", ".tiff"}
MVP_SHEETS = {"公司主体", "公司资料", "人员资料", "附件索引"}
FINANCIAL_SHEET = "财务报表"
SPECIALTY_LEDGER_SHEETS = {
    "银行账户": "bank_account",
    "保证金": "bid_bond",
    "绿证": "green_certificate",
    "科技成果": "technology_achievement",
    "ESG": "esg_report",
    "奖项": "award",
}
SUPPORTED_SHEETS = MVP_SHEETS | {FINANCIAL_SHEET} | set(SPECIALTY_LEDGER_SHEETS)
SUPPORTED_OWNER_TYPES = {"library_company", "company_profile", "person_profile"}
JSON_FIELDS = {"metadata_json", "profile_json", "statement_data"}
DATE_FIELDS = {"issued_on", "expires_on"}


@dataclass(frozen=True)
class CompanybaseIssue:
    severity: str
    sheet: str
    row: int | None
    message: str

    def to_dict(self) -> dict[str, Any]:
        return {"severity": self.severity, "sheet": self.sheet, "row": self.row, "message": self.message}


@dataclass
class CompanybaseReport:
    summary: dict[str, int] = field(default_factory=dict)
    issues: list[CompanybaseIssue] = field(default_factory=list)
    actions: dict[str, int] = field(default_factory=lambda: {"created": 0, "updated": 0, "skipped": 0})
    dry_run: bool = True

    @property
    def p0_count(self) -> int:
        return sum(1 for issue in self.issues if issue.severity == "P0")

    @property
    def p1_count(self) -> int:
        return sum(1 for issue in self.issues if issue.severity == "P1")

    def to_dict(self) -> dict[str, Any]:
        return {
            "summary": self.summary,
            "issues": [issue.to_dict() for issue in self.issues],
            "p0_count": self.p0_count,
            "p1_count": self.p1_count,
            "actions": self.actions,
            "dry_run": self.dry_run,
        }


class CompanybaseImportService:
    def __init__(self, *, files_root: Path | None = None) -> None:
        self.files_root = (files_root or Path(__file__).resolve().parents[3] / "companybase").resolve()

    def validate_workbook(self, workbook_path: Path) -> CompanybaseReport:
        tables = self._read_workbook(workbook_path)
        return self._validate_tables(tables)

    def import_workbook(self, conn: Connection, workbook_path: Path, *, dry_run: bool) -> CompanybaseReport:
        tables = self._read_workbook(workbook_path)
        report = self._validate_tables(tables)
        report.dry_run = dry_run
        if report.p0_count:
            return report
        if dry_run:
            report.actions = self._plan_actions(conn, tables)
            return report
        report.actions = self._apply_import(conn, tables)
        conn.commit()
        return report

    def create_backup_archive(self) -> Path:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        tmp = Path(tempfile.gettempdir()) / f"companybase_{stamp}.tar.gz"
        root = self.files_root
        with tarfile.open(tmp, "w:gz") as tar:
            for path in root.rglob("*"):
                rel = path.relative_to(root)
                if rel.parts and rel.parts[0] == "backups":
                    continue
                tar.add(path, arcname=Path("companybase") / rel)
        return tmp

    def _read_workbook(self, workbook_path: Path) -> dict[str, list[dict[str, str]]]:
        wb = load_workbook(workbook_path, read_only=True, data_only=True)
        tables: dict[str, list[dict[str, str]]] = {}
        for ws in wb.worksheets:
            if ws.title in {"说明", "字段字典"}:
                continue
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                tables[ws.title] = []
                continue
            header = [self._cell(value) for value in rows[0]]
            data: list[dict[str, str]] = []
            for row in rows[1:]:
                values = [self._cell(value) for value in row]
                if not any(values):
                    continue
                padded = values + [""] * max(0, len(header) - len(values))
                data.append(dict(zip(header, padded)))
            tables[ws.title] = data
        return tables

    def _validate_tables(self, tables: dict[str, list[dict[str, str]]]) -> CompanybaseReport:
        report = CompanybaseReport(summary={sheet: len(rows) for sheet, rows in tables.items()})
        company_keys = {row.get("company_key", "").strip() for row in tables.get("公司主体", []) if row.get("company_key", "").strip()}
        unique_by_owner = {
            "library_company": company_keys,
            "company_profile": {row.get("unique_key", "").strip() for row in tables.get("公司资料", []) if row.get("unique_key", "").strip()},
            "person_profile": {row.get("unique_key", "").strip() for row in tables.get("人员资料", []) if row.get("unique_key", "").strip()},
        }
        attachment_keys = {
            row.get("attachment_key", "").strip()
            for row in tables.get("附件索引", [])
            if row.get("attachment_key", "").strip()
        }

        for sheet, rows in tables.items():
            if sheet not in SUPPORTED_SHEETS:
                if rows:
                    report.issues.append(CompanybaseIssue("P1", sheet, None, "sheet not imported"))
                continue
            seen: set[str] = set()
            key_field = "company_key" if sheet == "公司主体" else "attachment_key" if sheet == "附件索引" else "unique_key"
            for index, row in enumerate(rows, start=2):
                key = row.get(key_field, "").strip()
                if not key:
                    report.issues.append(CompanybaseIssue("P0", sheet, index, f"{key_field} is required"))
                elif key in seen:
                    report.issues.append(CompanybaseIssue("P0", sheet, index, f"duplicate {key_field}: {key}"))
                seen.add(key)

                company_key = row.get("company_key", "").strip()
                if sheet != "公司主体":
                    if not company_key:
                        report.issues.append(CompanybaseIssue("P0", sheet, index, "company_key is required"))
                    elif company_key not in company_keys:
                        report.issues.append(CompanybaseIssue("P0", sheet, index, f"company_key not found: {company_key}"))

                for field_name in JSON_FIELDS:
                    raw = row.get(field_name, "").strip()
                    if raw and "__invalid_json__" in self._json(raw):
                        report.issues.append(CompanybaseIssue("P0", sheet, index, f"{field_name} must be a JSON object"))

                for field_name in DATE_FIELDS:
                    raw = row.get(field_name, "").strip()
                    if raw and not DATE_RE.match(raw):
                        report.issues.append(CompanybaseIssue("P0", sheet, index, f"{field_name} must be YYYY-MM-DD"))

                if sheet == "附件索引":
                    owner_type = row.get("owner_type", "").strip()
                    owner_key = row.get("owner_unique_key", "").strip()
                    rel = row.get("file_relative_path", "").strip()
                    if owner_type not in SUPPORTED_OWNER_TYPES:
                        report.issues.append(CompanybaseIssue("P1", sheet, index, f"owner_type not imported in MVP: {owner_type}"))
                    elif owner_key and owner_key not in unique_by_owner.get(owner_type, set()):
                        report.issues.append(CompanybaseIssue("P0", sheet, index, f"owner_unique_key not found for {owner_type}: {owner_key}"))
                    if not rel:
                        report.issues.append(CompanybaseIssue("P0", sheet, index, "file_relative_path is required"))
                    else:
                        path = (self.files_root / rel).resolve()
                        if self.files_root not in path.parents and path != self.files_root:
                            report.issues.append(CompanybaseIssue("P0", sheet, index, f"file path escapes companybase: {rel}"))
                        if path.suffix.lower() not in ALLOWED_SUFFIXES:
                            report.issues.append(CompanybaseIssue("P0", sheet, index, f"unsupported file extension: {rel}"))
                        if not path.is_file():
                            report.issues.append(CompanybaseIssue("P1", sheet, index, f"file does not exist yet: {rel}"))
                if sheet == FINANCIAL_SHEET:
                    fiscal_year = row.get("fiscal_year", row.get("year", "")).strip()
                    if not fiscal_year.isdigit():
                        report.issues.append(CompanybaseIssue("P0", sheet, index, "fiscal_year must be an integer"))
                    if not row.get("statement_type", "").strip():
                        report.issues.append(CompanybaseIssue("P0", sheet, index, "statement_type is required"))
                if sheet in SPECIALTY_LEDGER_SHEETS:
                    year = row.get("year", "").strip()
                    if year and not year.isdigit():
                        report.issues.append(CompanybaseIssue("P0", sheet, index, "year must be an integer"))
                    evidence_key = row.get("evidence_attachment_key", "").strip()
                    if evidence_key and evidence_key not in attachment_keys:
                        report.issues.append(CompanybaseIssue("P0", sheet, index, f"evidence_attachment_key not found: {evidence_key}"))
        return report

    def _plan_actions(self, conn: Connection, tables: dict[str, list[dict[str, str]]]) -> dict[str, int]:
        counts = {"created": 0, "updated": 0, "skipped": 0}
        for row in tables.get("公司主体", []):
            counts["updated" if self._find_library_company(conn, row.get("company_key", "")) else "created"] += 1
        for row in tables.get("公司资料", []):
            counts["updated" if self._find_by_import_key(conn, "company_profile", row.get("unique_key", "")) else "created"] += 1
        for row in tables.get("人员资料", []):
            counts["updated" if self._find_by_import_key(conn, "person_profile", row.get("unique_key", "")) else "created"] += 1
        for row in tables.get("附件索引", []):
            if row.get("owner_type", "").strip() not in SUPPORTED_OWNER_TYPES:
                counts["skipped"] += 1
            elif not (self.files_root / row.get("file_relative_path", "").strip()).is_file():
                counts["skipped"] += 1
            else:
                counts["updated" if self._find_by_import_key(conn, "evidence_asset", row.get("attachment_key", ""), json_column="metadata_json", key_name="attachment_key") else "created"] += 1
        for row in tables.get(FINANCIAL_SHEET, []):
            counts[
                "updated"
                if self._find_by_import_key(conn, "financial_statement", row.get("unique_key", ""), json_column="statement_data")
                else "created"
            ] += 1
        for sheet in SPECIALTY_LEDGER_SHEETS:
            for row in tables.get(sheet, []):
                counts[
                    "updated"
                    if self._find_by_import_key(conn, "business_specialty_ledger", row.get("unique_key", ""), json_column="metadata_json")
                    else "created"
                ] += 1
        return counts

    def _apply_import(self, conn: Connection, tables: dict[str, list[dict[str, str]]]) -> dict[str, int]:
        counts = {"created": 0, "updated": 0, "skipped": 0}
        company_ids: dict[str, UUID] = {}
        owner_ids: dict[tuple[str, str], UUID] = {}
        evidence_ids: dict[str, UUID] = {}

        for row in tables.get("公司主体", []):
            company_key = row.get("company_key", "").strip()
            existing = self._find_library_company(conn, company_key)
            metadata = self._with_import(row.get("metadata_json"), unique_key=company_key)
            fields = {
                "company_key": company_key,
                "company_name": row.get("company_name", "").strip(),
                "company_type": row.get("company_type", "").strip() or None,
                "enabled": row.get("enabled", "TRUE").strip().upper() != "FALSE",
                "metadata_json": metadata,
            }
            if existing:
                company_id = existing
                self._update_record(conn, "library_company", company_id, fields, json_fields={"metadata_json"})
                counts["updated"] += 1
            else:
                company_id = uuid4()
                self._insert_library_company(conn, company_id, fields)
                counts["created"] += 1
            company_ids[company_key] = company_id
            owner_ids[("library_company", company_key)] = company_id

        for row in tables.get("公司资料", []):
            unique_key = row.get("unique_key", "").strip()
            company_id = company_ids.get(row.get("company_key", "").strip()) or self._find_library_company(conn, row.get("company_key", ""))
            existing = self._find_by_import_key(conn, "company_profile", unique_key)
            fields = {
                "library_company_id": company_id,
                "company_name": row.get("company_name", "").strip(),
                "company_code": row.get("company_code", "").strip() or None,
                "unified_social_credit_code": row.get("unified_social_credit_code", "").strip() or None,
                "registered_address": row.get("registered_address", "").strip() or None,
                "contact_name": row.get("contact_name", "").strip() or None,
                "contact_phone": row.get("contact_phone", "").strip() or None,
                "contact_email": row.get("contact_email", "").strip() or None,
                "website": row.get("website", "").strip() or None,
                "registered_capital": row.get("registered_capital", "").strip() or None,
                "company_type": row.get("company_type", "").strip() or None,
                "business_scope": row.get("business_scope", "").strip() or None,
                "profile_json": self._with_import(row.get("profile_json"), unique_key=unique_key),
            }
            if existing:
                record_id = existing
                self._update_record(conn, "company_profile", record_id, fields, json_fields={"profile_json"})
                counts["updated"] += 1
            else:
                record_id = uuid4()
                self._insert_company_profile(conn, record_id, fields)
                counts["created"] += 1
            owner_ids[("company_profile", unique_key)] = record_id

        for row in tables.get("人员资料", []):
            unique_key = row.get("unique_key", "").strip()
            company_id = company_ids.get(row.get("company_key", "").strip()) or self._find_library_company(conn, row.get("company_key", ""))
            existing = self._find_by_import_key(conn, "person_profile", unique_key)
            fields = {
                "library_company_id": company_id,
                "full_name": row.get("full_name", "").strip(),
                "gender": row.get("gender", "").strip() or None,
                "age": self._int(row.get("age")),
                "education": row.get("education", "").strip() or None,
                "title": row.get("title", "").strip() or None,
                "role_name": row.get("role_name", "").strip() or None,
                "specialty": row.get("specialty", "").strip() or None,
                "years_experience": self._int(row.get("years_experience")),
                "phone": row.get("phone", "").strip() or None,
                "email": row.get("email", "").strip() or None,
                "resume_text": row.get("resume_text", "").strip() or None,
                "profile_json": self._with_import(row.get("profile_json"), unique_key=unique_key),
            }
            if existing:
                record_id = existing
                self._update_record(conn, "person_profile", record_id, fields, json_fields={"profile_json"})
                counts["updated"] += 1
            else:
                record_id = uuid4()
                self._insert_person(conn, record_id, fields)
                counts["created"] += 1
            owner_ids[("person_profile", unique_key)] = record_id

        for row in tables.get("附件索引", []):
            owner_type = row.get("owner_type", "").strip()
            if owner_type not in SUPPORTED_OWNER_TYPES:
                counts["skipped"] += 1
                continue
            rel = row.get("file_relative_path", "").strip()
            if not (self.files_root / rel).is_file():
                counts["skipped"] += 1
                continue
            unique_key = row.get("attachment_key", "").strip()
            existing = self._find_by_import_key(conn, "evidence_asset", unique_key, json_column="metadata_json", key_name="attachment_key")
            owner_key = row.get("owner_unique_key", "").strip()
            owner_id = owner_ids.get((owner_type, owner_key)) or self._find_owner_id(conn, owner_type, owner_key)
            company_id = company_ids.get(row.get("company_key", "").strip()) or self._find_library_company(conn, row.get("company_key", ""))
            fields = {
                "library_company_id": company_id,
                "owner_type": owner_type,
                "owner_id": owner_id,
                "asset_name": row.get("asset_name", "").strip() or unique_key,
                "asset_domain": row.get("asset_domain", "").strip() or "generic",
                "asset_category": row.get("asset_category", "").strip() or "supporting_document",
                "asset_type": row.get("asset_type", "").strip() or "supporting_document",
                "file_name": Path(rel).name,
                "file_path": str((self.files_root / rel).resolve()),
                "media_type": row.get("media_type", "").strip() or None,
                "issuer_name": row.get("issuer_name", "").strip() or None,
                "issued_on": self._date(row.get("issued_on")),
                "expires_on": self._date(row.get("expires_on")),
                "metadata_json": self._attachment_metadata(row),
                "sort_order": self._int(row.get("sort_order")) or 0,
            }
            if existing:
                self._update_record(conn, "evidence_asset", existing, fields, json_fields={"metadata_json"})
                evidence_id = existing
                counts["updated"] += 1
            else:
                evidence_id = uuid4()
                self._insert_evidence(conn, evidence_id, fields)
                counts["created"] += 1
            evidence_ids[unique_key] = evidence_id
        for row in tables.get(FINANCIAL_SHEET, []):
            unique_key = row.get("unique_key", "").strip()
            existing = self._find_by_import_key(conn, "financial_statement", unique_key, json_column="statement_data")
            fields = {
                "library_company_id": company_ids.get(row.get("company_key", "").strip()) or self._find_library_company(conn, row.get("company_key", "")),
                "fiscal_year": self._int(row.get("fiscal_year") or row.get("year")),
                "statement_type": row.get("statement_type", "").strip(),
                "statement_data": self._financial_statement_data(row, unique_key=unique_key),
                "source_note": row.get("source_note", "").strip() or None,
            }
            if existing:
                self._update_record(conn, "financial_statement", existing, fields, json_fields={"statement_data"})
                counts["updated"] += 1
            else:
                self._insert_financial_statement(conn, uuid4(), fields)
                counts["created"] += 1
        for sheet, ledger_type in SPECIALTY_LEDGER_SHEETS.items():
            for row in tables.get(sheet, []):
                unique_key = row.get("unique_key", "").strip()
                evidence_key = row.get("evidence_attachment_key", "").strip()
                evidence_id = evidence_ids.get(evidence_key)
                if evidence_key and evidence_id is None:
                    evidence_id = self._find_by_import_key(
                        conn,
                        "evidence_asset",
                        evidence_key,
                        json_column="metadata_json",
                        key_name="attachment_key",
                    )
                existing = self._find_by_import_key(
                    conn,
                    "business_specialty_ledger",
                    unique_key,
                    json_column="metadata_json",
                )
                fields = {
                    "library_company_id": company_ids.get(row.get("company_key", "").strip()) or self._find_library_company(conn, row.get("company_key", "")),
                    "company_key": row.get("company_key", "").strip() or None,
                    "ledger_type": ledger_type,
                    "year": self._int(row.get("year")),
                    "evidence_asset_id": evidence_id,
                    "metadata_json": self._specialty_ledger_metadata(row, unique_key=unique_key, source_sheet=sheet, ledger_type=ledger_type),
                }
                if existing:
                    self._update_record(conn, "business_specialty_ledger", existing, fields, json_fields={"metadata_json"})
                    counts["updated"] += 1
                else:
                    self._insert_business_specialty_ledger(conn, uuid4(), fields)
                    counts["created"] += 1
        return counts

    @staticmethod
    def _cell(value: Any) -> str:
        if value is None:
            return ""
        return str(value).strip()

    @staticmethod
    def _json(raw: str | None) -> dict[str, Any]:
        if not raw or not str(raw).strip():
            return {}
        try:
            value = json.loads(str(raw))
        except json.JSONDecodeError:
            return {"__invalid_json__": str(raw)}
        return value if isinstance(value, dict) else {"__invalid_json__": str(raw)}

    def _with_import(self, raw: str | None, *, unique_key: str) -> dict[str, Any]:
        data = self._json(raw)
        data["import"] = {**dict(data.get("import") or {}), "source": "companybase", "unique_key": unique_key}
        return data

    def _attachment_metadata(self, row: dict[str, str]) -> dict[str, Any]:
        data = self._json(row.get("metadata_json"))
        data["import"] = {
            **dict(data.get("import") or {}),
            "source": "companybase",
            "attachment_key": row.get("attachment_key", "").strip(),
            "owner_unique_key": row.get("owner_unique_key", "").strip(),
        }
        data["blind_bid"] = {
            "sensitive": row.get("is_blind_sensitive", "").strip().upper() == "TRUE",
            "redaction_note": row.get("redaction_note", "").strip(),
        }
        return data

    def _financial_statement_data(self, row: dict[str, str], *, unique_key: str) -> dict[str, Any]:
        data = self._with_import(row.get("statement_data"), unique_key=unique_key)
        data["import"]["source_sheet"] = FINANCIAL_SHEET
        return data

    def _specialty_ledger_metadata(
        self,
        row: dict[str, str],
        *,
        unique_key: str,
        source_sheet: str,
        ledger_type: str,
    ) -> dict[str, Any]:
        data = self._with_import(row.get("metadata_json"), unique_key=unique_key)
        data["import"]["source_sheet"] = source_sheet
        data["ledger_type"] = ledger_type
        evidence_key = row.get("evidence_attachment_key", "").strip()
        if evidence_key:
            data["evidence_attachment_key"] = evidence_key
        return data

    @staticmethod
    def _int(value: Any) -> int | None:
        raw = str(value or "").strip()
        return int(raw) if raw.isdigit() else None

    @staticmethod
    def _date(value: Any) -> date | None:
        raw = str(value or "").strip()
        return date.fromisoformat(raw) if DATE_RE.match(raw) else None

    def _find_library_company(self, conn: Connection, company_key: str) -> UUID | None:
        with conn.cursor() as cur:
            row = cur.execute("SELECT id FROM library_company WHERE company_key = %s", (company_key.strip(),)).fetchone()
        return row[0] if row else None

    def _find_by_import_key(self, conn: Connection, table: str, key: str, *, json_column: str = "profile_json", key_name: str = "unique_key") -> UUID | None:
        with conn.cursor() as cur:
            row = cur.execute(
                f"SELECT id FROM {table} WHERE {json_column}->'import'->>%s = %s LIMIT 1",
                (key_name, key.strip()),
            ).fetchone()
        return row[0] if row else None

    def _find_owner_id(self, conn: Connection, owner_type: str, owner_key: str) -> UUID | None:
        if owner_type == "library_company":
            return self._find_library_company(conn, owner_key)
        table = "company_profile" if owner_type == "company_profile" else "person_profile"
        return self._find_by_import_key(conn, table, owner_key)

    def _insert_library_company(self, conn: Connection, record_id: UUID, fields: dict[str, Any]) -> None:
        conn.execute(
            """
            INSERT INTO library_company (id, company_key, company_name, company_type, enabled, metadata_json)
            VALUES (%s, %s, %s, %s, %s, %s::jsonb)
            """,
            (record_id, fields["company_key"], fields["company_name"], fields["company_type"], fields["enabled"], json.dumps(fields["metadata_json"], ensure_ascii=False)),
        )

    def _insert_company_profile(self, conn: Connection, record_id: UUID, fields: dict[str, Any]) -> None:
        conn.execute(
            """
            INSERT INTO company_profile (
              id, library_company_id, company_name, company_code, unified_social_credit_code, registered_address,
              contact_name, contact_phone, contact_email, website, registered_capital, company_type, business_scope, profile_json
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s::jsonb)
            """,
            (record_id, fields["library_company_id"], fields["company_name"], fields["company_code"], fields["unified_social_credit_code"], fields["registered_address"], fields["contact_name"], fields["contact_phone"], fields["contact_email"], fields["website"], fields["registered_capital"], fields["company_type"], fields["business_scope"], json.dumps(fields["profile_json"], ensure_ascii=False)),
        )

    def _insert_person(self, conn: Connection, record_id: UUID, fields: dict[str, Any]) -> None:
        conn.execute(
            """
            INSERT INTO person_profile (
              id, library_company_id, full_name, gender, age, education, title, role_name, specialty,
              years_experience, phone, email, resume_text, profile_json
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s::jsonb)
            """,
            (record_id, fields["library_company_id"], fields["full_name"], fields["gender"], fields["age"], fields["education"], fields["title"], fields["role_name"], fields["specialty"], fields["years_experience"], fields["phone"], fields["email"], fields["resume_text"], json.dumps(fields["profile_json"], ensure_ascii=False)),
        )

    def _insert_evidence(self, conn: Connection, record_id: UUID, fields: dict[str, Any]) -> None:
        conn.execute(
            """
            INSERT INTO evidence_asset (
              id, library_company_id, owner_type, owner_id, asset_name, asset_domain, asset_category,
              asset_type, file_name, file_path, media_type, issuer_name, issued_on, expires_on, metadata_json, sort_order
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s::jsonb, %s)
            """,
            (record_id, fields["library_company_id"], fields["owner_type"], fields["owner_id"], fields["asset_name"], fields["asset_domain"], fields["asset_category"], fields["asset_type"], fields["file_name"], fields["file_path"], fields["media_type"], fields["issuer_name"], fields["issued_on"], fields["expires_on"], json.dumps(fields["metadata_json"], ensure_ascii=False), fields["sort_order"]),
        )

    def _insert_financial_statement(self, conn: Connection, record_id: UUID, fields: dict[str, Any]) -> None:
        conn.execute(
            """
            INSERT INTO financial_statement (
              id, library_company_id, fiscal_year, statement_type, statement_data, source_note
            ) VALUES (%s, %s, %s, %s, %s::jsonb, %s)
            """,
            (
                record_id,
                fields["library_company_id"],
                fields["fiscal_year"],
                fields["statement_type"],
                json.dumps(fields["statement_data"], ensure_ascii=False),
                fields["source_note"],
            ),
        )

    def _insert_business_specialty_ledger(self, conn: Connection, record_id: UUID, fields: dict[str, Any]) -> None:
        conn.execute(
            """
            INSERT INTO business_specialty_ledger (
              id, library_company_id, company_key, ledger_type, year, evidence_asset_id, metadata_json
            ) VALUES (%s, %s, %s, %s, %s, %s, %s::jsonb)
            """,
            (
                record_id,
                fields["library_company_id"],
                fields["company_key"],
                fields["ledger_type"],
                fields["year"],
                fields["evidence_asset_id"],
                json.dumps(fields["metadata_json"], ensure_ascii=False),
            ),
        )

    def _update_record(self, conn: Connection, table: str, record_id: UUID, fields: dict[str, Any], *, json_fields: set[str]) -> None:
        sets: list[str] = []
        values: list[Any] = []
        for key, value in fields.items():
            if key in json_fields:
                sets.append(f"{key} = %s::jsonb")
                values.append(json.dumps(value or {}, ensure_ascii=False))
            else:
                sets.append(f"{key} = %s")
                values.append(value)
        sets.append("updated_at = now()")
        values.append(record_id)
        conn.execute(f"UPDATE {table} SET {', '.join(sets)} WHERE id = %s", values)
