from __future__ import annotations

import io
from collections import defaultdict
from typing import Any
from uuid import UUID

from openpyxl import Workbook
from openpyxl.styles import Font
from psycopg import Connection
from psycopg.rows import dict_row


_ASSET_LABELS = {
    "vehicle": "车辆",
    "machine": "施工机械",
    "tool": "施工工器具",
    "safety": "安全设施设备及器具",
}

_COLUMNS = {
    "vehicle": [
        ("序号", "index"),
        ("设备名称", "name"),
        ("规格型号", "spec_model"),
        ("车牌/编号", "serial_no"),
        ("数量", "quantity_unit"),
        ("所有权", "ownership"),
        ("车辆类型", "extra:vehicle_type"),
        ("技术状况", "technical_condition"),
        ("用途", "intended_role"),
    ],
    "machine": [
        ("序号", "index"),
        ("设备名称", "name"),
        ("规格型号", "spec_model"),
        ("出厂编号", "serial_no"),
        ("数量", "quantity_unit"),
        ("所有权", "ownership"),
        ("机械类别", "extra:machine_category"),
        ("容量", "extra:capacity"),
        ("用途", "intended_role"),
    ],
    "tool": [
        ("序号", "index"),
        ("设备名称", "name"),
        ("规格型号", "spec_model"),
        ("出厂编号", "serial_no"),
        ("数量", "quantity_unit"),
        ("所有权", "ownership"),
        ("工器具类别", "extra:tool_category"),
        ("电压等级", "extra:voltage_level"),
        ("用途", "intended_role"),
    ],
    "safety": [
        ("序号", "index"),
        ("设备名称", "name"),
        ("规格型号", "spec_model"),
        ("数量", "quantity_unit"),
        ("所有权", "ownership"),
        ("安全类别", "extra:safety_category"),
        ("防护标准", "extra:protection_standard"),
        ("适用工种", "extra:applicable_work"),
        ("用途", "intended_role"),
    ],
}


def _ownership_label(value: str | None) -> str:
    if value == "self":
        return "自有"
    if value == "leased":
        return "租赁"
    if value == "third_party":
        return "第三方"
    return value or "—"


class EquipmentTableRenderer:
    def render_equipment_preview(self, conn: Connection, *, project_id: UUID) -> dict[str, list[dict[str, str]]]:
        grouped = self._load_confirmed(conn, project_id=project_id)
        result: dict[str, list[dict[str, str]]] = {}
        for asset_type in _ASSET_LABELS:
            rows = grouped.get(asset_type, [])
            result[asset_type] = [self._row_to_preview_row(index + 1, row) for index, row in enumerate(rows)]
        return result

    def render_attachment_xlsx(self, conn: Connection, *, project_id: UUID) -> bytes:
        grouped = self._load_confirmed(conn, project_id=project_id)
        workbook = Workbook()
        first_sheet = True
        for asset_type, sheet_name in _ASSET_LABELS.items():
            sheet = workbook.active if first_sheet else workbook.create_sheet()
            first_sheet = False
            sheet.title = sheet_name
            columns = _COLUMNS[asset_type]
            for column_index, (label, _source) in enumerate(columns, start=1):
                cell = sheet.cell(row=1, column=column_index, value=label)
                cell.font = Font(bold=True)
            rows = grouped.get(asset_type, [])
            if not rows:
                sheet.cell(row=2, column=1, value="无")
            else:
                for row_index, row in enumerate(rows, start=2):
                    preview = self._row_to_preview_row(row_index - 1, row)
                    for column_index, (label, _source) in enumerate(columns, start=1):
                        sheet.cell(row=row_index, column=column_index, value=preview[label])
            sheet.freeze_panes = "A2"
            for column in sheet.columns:
                max_length = max(len(str(cell.value or "")) for cell in column)
                sheet.column_dimensions[column[0].column_letter].width = min(max(max_length + 2, 10), 28)
        buffer = io.BytesIO()
        workbook.save(buffer)
        return buffer.getvalue()

    def _load_confirmed(self, conn: Connection, *, project_id: UUID) -> dict[str, list[dict[str, Any]]]:
        with conn.cursor(row_factory=dict_row) as cur:
            rows = cur.execute(
                """
                SELECT asset_type, intended_role, snapshot_json, display_order, created_at
                FROM project_equipment_selection
                WHERE project_id = %s AND confirmed = TRUE
                ORDER BY asset_type, display_order, created_at
                """,
                (project_id,),
            ).fetchall()
        grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for row in rows:
            snapshot = dict(row["snapshot_json"] or {})
            snapshot["asset_type"] = row["asset_type"]
            snapshot["intended_role"] = row["intended_role"]
            grouped[row["asset_type"]].append(snapshot)
        return grouped

    def _row_to_preview_row(self, index: int, row: dict[str, Any]) -> dict[str, str]:
        asset_type = str(row.get("asset_type") or "")
        data: dict[str, str] = {}
        for label, source in _COLUMNS[asset_type]:
            data[label] = self._resolve_value(index, row, source)
        return data

    def _resolve_value(self, index: int, row: dict[str, Any], source: str) -> str:
        if source == "index":
            return str(index)
        if source == "quantity_unit":
            quantity = row.get("quantity")
            unit = row.get("unit")
            if quantity in (None, ""):
                return "—"
            return f"{quantity}{unit or ''}"
        if source == "ownership":
            return _ownership_label(row.get("ownership"))
        if source.startswith("extra:"):
            extras = dict(row.get("extras") or {})
            value = extras.get(source.split(":", 1)[1])
            return str(value) if value not in (None, "") else "—"
        value = row.get(source)
        return str(value) if value not in (None, "") else "—"
