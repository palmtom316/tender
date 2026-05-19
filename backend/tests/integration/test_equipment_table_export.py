from __future__ import annotations

import io
import os
from uuid import uuid4

import psycopg
import pytest
from openpyxl import load_workbook

from tender_backend.services.export_service.equipment_table_renderer import EquipmentTableRenderer


def _db_url() -> str | None:
    return os.environ.get("DATABASE_URL")


@pytest.mark.skipif(not _db_url(), reason="DATABASE_URL not set")
def test_equipment_table_renderer_preview_and_xlsx() -> None:
    renderer = EquipmentTableRenderer()
    project_id = uuid4()
    with psycopg.connect(_db_url(), autocommit=True) as conn:
        library_company_id = uuid4()
        asset_id = uuid4()
        conn.execute(
            "INSERT INTO library_company (id, company_key, company_name) VALUES (%s, %s, %s)",
            (library_company_id, f"l-{library_company_id.hex[:8]}", "重庆示例电力工程有限责任公司"),
        )
        conn.execute("INSERT INTO project (id, name) VALUES (%s, %s)", (project_id, "测试项目"))
        conn.execute(
            """
            INSERT INTO company_asset (
              id, library_company_id, asset_type, name, spec_model, serial_no, manufacturer,
              quantity, unit, ownership, extras
            ) VALUES (%s, %s, 'vehicle', '斗臂车', 'DFL5160', '渝A12345', '东风', 1, '辆', 'self', '{"vehicle_type":"aerial_bucket"}'::jsonb)
            """,
            (asset_id, library_company_id),
        )
        conn.execute(
            """
            INSERT INTO project_equipment_selection (
              id, project_id, asset_id, asset_type, intended_role, snapshot_json, confirmed, confirmed_at
            ) VALUES (%s, %s, %s, 'vehicle', '配电主线', %s::jsonb, TRUE, now())
            """,
            (
                uuid4(),
                project_id,
                asset_id,
                '{"name":"斗臂车","spec_model":"DFL5160","serial_no":"渝A12345","manufacturer":"东风","quantity":"1","unit":"辆","ownership":"self","technical_condition":"良好","extras":{"vehicle_type":"aerial_bucket"}}',
            ),
        )

        preview = renderer.render_equipment_preview(conn, project_id=project_id)
        assert len(preview["vehicle"]) == 1
        assert preview["vehicle"][0]["设备名称"] == "斗臂车"
        assert preview["vehicle"][0]["用途"] == "配电主线"

        data = renderer.render_attachment_xlsx(conn, project_id=project_id)
    workbook = load_workbook(io.BytesIO(data))
    assert sorted(workbook.sheetnames) == sorted(["车辆", "施工机械", "施工工器具", "安全设施设备及器具"])
    assert workbook["车辆"]["A2"].value == "1"
    assert workbook["车辆"]["B2"].value == "斗臂车"
