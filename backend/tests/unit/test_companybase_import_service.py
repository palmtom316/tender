from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from tender_backend.services.companybase_import_service import CompanybaseImportService


def _write_workbook(path: Path) -> None:
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    sheets = {
        "公司主体": [
            ["company_key", "company_name", "company_type", "enabled", "metadata_json"],
            ["main_company", "某某电力工程有限公司", "bidder", "TRUE", '{"source":"test"}'],
        ],
        "公司资料": [
            ["unique_key", "company_key", "company_name", "company_code", "unified_social_credit_code", "registered_address", "contact_name", "contact_phone", "contact_email", "website", "registered_capital", "company_type", "business_scope", "profile_json"],
            ["profile_main", "main_company", "某某电力工程有限公司", "MC-001", "91500000000000000X", "重庆市某区", "联系人", "13800000000", "contact@example.com", "", "5000万元", "有限责任公司", "电力工程施工", '{"blind_bid_alias":"投标人"}'],
        ],
        "人员资料": [
            ["unique_key", "company_key", "full_name", "gender", "age", "education", "title", "role_name", "specialty", "years_experience", "phone", "email", "resume_text", "profile_json"],
            ["person_pm", "main_company", "张三", "男", "38", "本科", "工程师", "项目经理", "机电工程", "12", "13800000001", "pm@example.com", "配网施工管理", '{"blind_bid_alias":"项目经理"}'],
        ],
        "附件索引": [
            ["attachment_key", "company_key", "owner_type", "owner_unique_key", "asset_name", "asset_domain", "asset_category", "asset_type", "file_relative_path", "media_type", "issuer_name", "issued_on", "expires_on", "is_blind_sensitive", "redaction_note", "sort_order", "metadata_json"],
            ["att_person_pm", "main_company", "person_profile", "person_pm", "项目经理证书", "personnel", "practice_certificate", "constructor_certificate", "files/personnel/pm.pdf", "application/pdf", "主管部门", "2024-01-01", "2028-01-01", "TRUE", "含姓名", "10", '{"core_slot":"practice"}'],
        ],
        "财务报表": [
            ["unique_key", "company_key", "fiscal_year", "statement_type", "statement_data", "source_note"],
            ["finance_2025", "main_company", "2025", "annual_report", '{"revenue":1000}', "2025年审计报告"],
        ],
        "银行账户": [
            ["unique_key", "company_key", "year", "evidence_attachment_key", "metadata_json"],
            ["bank_main", "main_company", "2026", "att_person_pm", '{"bank_name":"开户银行"}'],
        ],
        "保证金": [
            ["unique_key", "company_key", "year", "metadata_json"],
            ["bond_2026", "main_company", "2026", '{"amount":10000}'],
        ],
        "绿证": [
            ["unique_key", "company_key", "year", "metadata_json"],
            ["green_2025", "main_company", "2025", '{"certificate_no":"GC-001"}'],
        ],
        "科技成果": [
            ["unique_key", "company_key", "year", "metadata_json"],
            ["tech_2025", "main_company", "2025", '{"achievement_name":"配网成果"}'],
        ],
        "ESG": [
            ["unique_key", "company_key", "year", "metadata_json"],
            ["esg_2025", "main_company", "2025", '{"report_name":"ESG报告"}'],
        ],
        "奖项": [
            ["unique_key", "company_key", "year", "metadata_json"],
            ["award_2025", "main_company", "2025", '{"award_name":"质量奖"}'],
        ],
    }
    for title, rows in sheets.items():
        ws = wb.create_sheet(title)
        for row in rows:
            ws.append(row)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def test_parse_workbook_counts_mvp_rows_and_warns_missing_attachment(tmp_path: Path) -> None:
    workbook = tmp_path / "companybase_master.xlsx"
    _write_workbook(workbook)

    report = CompanybaseImportService(files_root=tmp_path).validate_workbook(workbook)

    assert report.summary["公司主体"] == 1
    assert report.summary["公司资料"] == 1
    assert report.summary["人员资料"] == 1
    assert report.summary["附件索引"] == 1
    assert report.summary["财务报表"] == 1
    assert report.summary["银行账户"] == 1
    assert report.summary["保证金"] == 1
    assert report.summary["绿证"] == 1
    assert report.summary["科技成果"] == 1
    assert report.summary["ESG"] == 1
    assert report.summary["奖项"] == 1
    assert report.p0_count == 0
    assert any(issue.severity == "P1" and "file does not exist" in issue.message for issue in report.issues)
    assert not any("not imported in MVP" in issue.message for issue in report.issues)


def test_parse_workbook_reports_duplicate_unique_key(tmp_path: Path) -> None:
    workbook = tmp_path / "companybase_master.xlsx"
    _write_workbook(workbook)
    wb = Workbook()
    ws = wb.active
    ws.title = "人员资料"
    ws.append(["unique_key", "company_key", "full_name"])
    ws.append(["dup", "missing_company", "张三"])
    ws.append(["dup", "missing_company", "李四"])
    wb.save(workbook)

    report = CompanybaseImportService(files_root=tmp_path).validate_workbook(workbook)

    assert report.p0_count >= 2
    assert any("duplicate unique_key" in issue.message for issue in report.issues)
    assert any("company_key not found" in issue.message for issue in report.issues)


def test_parse_workbook_validates_new_sheet_json_year_and_attachment_reference(tmp_path: Path) -> None:
    workbook = tmp_path / "companybase_master.xlsx"
    _write_workbook(workbook)
    wb = Workbook()
    ws = wb.active
    ws.title = "银行账户"
    ws.append(["unique_key", "company_key", "year", "evidence_attachment_key", "metadata_json"])
    ws.append(["bank_bad", "main_company", "20X6", "missing_attachment", '{"bank_name":'])
    wb.create_sheet("公司主体").append(["company_key", "company_name"])
    wb["公司主体"].append(["main_company", "某某电力工程有限公司"])
    wb.save(workbook)

    report = CompanybaseImportService(files_root=tmp_path).validate_workbook(workbook)

    assert report.p0_count == 3
    assert any("year must be an integer" in issue.message for issue in report.issues)
    assert any("metadata_json must be a JSON object" in issue.message for issue in report.issues)
    assert any("evidence_attachment_key not found" in issue.message for issue in report.issues)
